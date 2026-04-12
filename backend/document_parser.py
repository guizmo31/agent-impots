"""
Module de parsing de documents fiscaux.
Supporte : PDF, images (OCR), Excel, CSV, Word, TXT.
"""
import os
from pathlib import Path


class DocumentParser:
    """Parse différents types de documents et en extrait le texte."""

    def parse(self, file_path: str) -> dict | None:
        """Parse un document et retourne ses métadonnées + contenu texte."""
        path = Path(file_path)
        if not path.exists():
            return None

        ext = path.suffix.lower()
        content = ""

        try:
            if ext == ".pdf":
                content = self._parse_pdf(path)
            elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
                content = self._parse_image(path)
            elif ext in (".xlsx", ".xls"):
                content = self._parse_excel(path)
            elif ext == ".csv":
                content = self._parse_csv(path)
            elif ext == ".docx":
                content = self._parse_docx(path)
            elif ext == ".txt":
                content = self._parse_txt(path)
            else:
                return None
        except Exception as e:
            content = f"[Erreur de lecture : {str(e)}]"

        if not content or not content.strip():
            return None

        return {
            "filename": path.name,
            "filepath": str(path),
            "extension": ext,
            "size_bytes": path.stat().st_size,
            "content": content.strip(),
        }

    def _parse_pdf(self, path: Path) -> str:
        """Extrait le texte d'un PDF avec PyMuPDF."""
        import fitz  # PyMuPDF

        text_parts = []
        with fitz.open(str(path)) as doc:
            for page_num, page in enumerate(doc, 1):
                page_text = page.get_text()
                if page_text.strip():
                    text_parts.append(f"[Page {page_num}]\n{page_text}")
                else:
                    # Si pas de texte, tenter l'OCR sur le rendu de la page
                    pix = page.get_pixmap(dpi=200)
                    img_data = pix.tobytes("png")
                    ocr_text = self._ocr_bytes(img_data)
                    if ocr_text.strip():
                        text_parts.append(f"[Page {page_num} - OCR]\n{ocr_text}")

        return "\n\n".join(text_parts)

    def _parse_image(self, path: Path) -> str:
        """Extrait le texte d'une image par OCR."""
        try:
            import pytesseract
            from PIL import Image

            img = Image.open(str(path))
            text = pytesseract.image_to_string(img, lang="fra")
            return text
        except ImportError:
            return self._parse_image_fallback(path)
        except Exception as e:
            return f"[OCR non disponible : {str(e)}. Installez Tesseract OCR pour activer cette fonctionnalité.]"

    def _parse_image_fallback(self, path: Path) -> str:
        """Fallback si pytesseract n'est pas installé."""
        return (
            f"[Image détectée : {path.name}. "
            "L'OCR n'est pas disponible. Installez Tesseract OCR : "
            "https://github.com/UB-Mannheim/tesseract/wiki pour Windows]"
        )

    def _ocr_bytes(self, img_bytes: bytes) -> str:
        """OCR sur des bytes d'image."""
        try:
            import pytesseract
            from PIL import Image
            import io

            img = Image.open(io.BytesIO(img_bytes))
            return pytesseract.image_to_string(img, lang="fra")
        except Exception:
            return ""

    def _parse_excel(self, path: Path) -> str:
        """Extrait le contenu d'un fichier Excel."""
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"[Feuille: {sheet_name}]\n" + "\n".join(rows))

        wb.close()
        return "\n\n".join(parts)

    def _parse_csv(self, path: Path) -> str:
        """Extrait le contenu d'un fichier CSV."""
        import csv

        rows = []
        # Essayer différents encodages
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(str(path), "r", encoding=encoding) as f:
                    reader = csv.reader(f, delimiter=";")
                    for row in reader:
                        rows.append(" | ".join(row))
                break
            except UnicodeDecodeError:
                continue

        if not rows:
            # Fallback avec délimiteur virgule
            with open(str(path), "r", encoding="latin-1") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(" | ".join(row))

        return "\n".join(rows)

    def _parse_docx(self, path: Path) -> str:
        """Extrait le texte d'un fichier Word."""
        from docx import Document

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _parse_txt(self, path: Path) -> str:
        """Lit un fichier texte."""
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_text(encoding="latin-1", errors="replace")
