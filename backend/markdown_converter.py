"""
Convertisseur universel de documents en Markdown structure.

Pipeline : Fichier brut -> Markdown propre -> stocke sur disque

Le markdown intermediaire :
1. Est lisible par l'utilisateur (verification du contenu)
2. Est envoye au LLM au lieu du texte brut (plus propre = plus rapide)
3. Sert de cache (pas besoin de re-parser le fichier original)

Formats supportes : PDF, images (OCR), Excel, CSV, Word, TXT
"""
import csv
import io
import re
from datetime import datetime
from pathlib import Path


class MarkdownConverter:
    """Convertit n'importe quel document en fichier Markdown structure."""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.md_dir = self.output_dir / "markdown"
        self.md_dir.mkdir(parents=True, exist_ok=True)

    def convert(self, file_path: str) -> dict | None:
        """Convertit un fichier en markdown et le sauvegarde sur disque.

        Returns:
            {"filename", "md_path", "md_filename", "content", "size_bytes", "extension"}
            ou None si le fichier n'est pas lisible.
        """
        path = Path(file_path)
        if not path.exists():
            return None

        ext = path.suffix.lower()
        md_filename = path.stem + ".md"
        md_path = self.md_dir / md_filename

        # Cache : reutiliser le markdown existant si le fichier source n'a pas change
        if md_path.exists():
            source_mtime = path.stat().st_mtime
            md_mtime = md_path.stat().st_mtime
            if md_mtime >= source_mtime:
                # Le markdown est plus recent que le fichier source -> cache valide
                md_content = md_path.read_text(encoding="utf-8")
                if md_content.strip():
                    print(f"[MD] Cache hit : {path.name} -> {md_filename}")
                    return {
                        "filename": path.name,
                        "md_path": str(md_path),
                        "md_filename": md_filename,
                        "content": md_content,
                        "size_bytes": path.stat().st_size,
                        "extension": ext,
                        "cached": True,
                    }

        md_content = ""

        try:
            if ext == ".pdf":
                md_content = self._pdf_to_md(path)
            elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
                md_content = self._image_to_md(path)
            elif ext in (".xlsx", ".xls"):
                md_content = self._excel_to_md(path)
            elif ext == ".csv":
                md_content = self._csv_to_md(path)
            elif ext == ".docx":
                md_content = self._docx_to_md(path)
            elif ext == ".txt":
                md_content = self._txt_to_md(path)
            else:
                return None
        except Exception as e:
            md_content = f"# {path.name}\n\n**Erreur de lecture** : {str(e)}\n"

        if not md_content or not md_content.strip():
            return None

        # Ajouter un en-tete
        header = (
            f"<!-- Source : {path.name} -->\n"
            f"<!-- Converti le : {datetime.now().strftime('%d/%m/%Y %H:%M')} -->\n"
            f"<!-- Taille originale : {path.stat().st_size:,} octets -->\n\n"
        )
        md_content = header + md_content

        # Sauvegarder le markdown sur disque
        md_path.write_text(md_content, encoding="utf-8")
        print(f"[MD] Converti : {path.name} -> {md_filename} ({len(md_content)} chars)")

        return {
            "filename": path.name,
            "md_path": str(md_path),
            "md_filename": md_filename,
            "content": md_content,
            "size_bytes": path.stat().st_size,
            "extension": ext,
            "cached": False,
        }

    def get_all_markdowns(self) -> list[dict]:
        """Liste tous les markdowns generes."""
        results = []
        for f in sorted(self.md_dir.glob("*.md")):
            content = f.read_text(encoding="utf-8")
            # Extraire le nom source depuis le commentaire
            source_match = re.search(r"<!-- Source : (.+?) -->", content)
            source = source_match.group(1) if source_match else f.stem
            results.append({
                "md_filename": f.name,
                "source_filename": source,
                "size": len(content),
                "preview": content[:200],
            })
        return results

    # ------------------------------------------------------------------
    # Convertisseurs par format
    # ------------------------------------------------------------------

    def _pdf_to_md(self, path: Path) -> str:
        """PDF -> Markdown avec structure par pages."""
        import fitz  # PyMuPDF

        parts = [f"# {path.name}\n"]

        with fitz.open(str(path)) as doc:
            for page_num, page in enumerate(doc, 1):
                page_text = page.get_text()

                if not page_text.strip():
                    # Page sans texte -> OCR
                    pix = page.get_pixmap(dpi=200)
                    img_data = pix.tobytes("png")
                    page_text = self._ocr_bytes(img_data)
                    if page_text.strip():
                        parts.append(f"## Page {page_num} (OCR)\n")
                    else:
                        parts.append(f"## Page {page_num}\n\n*Page vide ou illisible*\n")
                        continue
                else:
                    parts.append(f"## Page {page_num}\n")

                # Nettoyer et structurer le texte
                cleaned = self._clean_text(page_text)
                # Detecter les lignes qui ressemblent a des montants (tableaux)
                structured = self._structure_fiscal_text(cleaned)
                parts.append(structured + "\n")

        return "\n".join(parts)

    def _image_to_md(self, path: Path) -> str:
        """Image -> Markdown via OCR."""
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(str(path))
            text = pytesseract.image_to_string(img, lang="fra")
            cleaned = self._clean_text(text)
            return f"# {path.name}\n\n*Document numerise (OCR)*\n\n{cleaned}\n"
        except ImportError:
            return f"# {path.name}\n\n*OCR non disponible (installez Tesseract)*\n"
        except Exception as e:
            return f"# {path.name}\n\n*Erreur OCR : {e}*\n"

    def _excel_to_md(self, path: Path) -> str:
        """Excel -> Markdown avec tableaux."""
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = [f"# {path.name}\n"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Feuille : {sheet_name}\n")

            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)

            if rows:
                # Construire un tableau markdown
                if len(rows) > 1:
                    header = rows[0]
                    parts.append("| " + " | ".join(header) + " |")
                    parts.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in rows[1:]:
                        # Aligner le nombre de colonnes
                        while len(row) < len(header):
                            row.append("")
                        parts.append("| " + " | ".join(row[:len(header)]) + " |")
                else:
                    parts.append("| " + " | ".join(rows[0]) + " |")
                parts.append("")
            else:
                parts.append("*Feuille vide*\n")

        wb.close()
        return "\n".join(parts)

    def _csv_to_md(self, path: Path) -> str:
        """CSV -> Markdown tableau."""
        rows = []
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(str(path), "r", encoding=encoding) as f:
                    # Detecter le delimiteur
                    sample = f.read(2000)
                    f.seek(0)
                    delimiter = ";" if sample.count(";") > sample.count(",") else ","
                    reader = csv.reader(f, delimiter=delimiter)
                    for row in reader:
                        if any(row):
                            rows.append(row)
                break
            except UnicodeDecodeError:
                continue

        if not rows:
            return f"# {path.name}\n\n*Fichier CSV vide ou illisible*\n"

        parts = [f"# {path.name}\n"]
        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:50]:  # Limiter a 50 lignes
            while len(row) < len(header):
                row.append("")
            parts.append("| " + " | ".join(row[:len(header)]) + " |")

        if len(rows) > 51:
            parts.append(f"\n*... et {len(rows) - 51} lignes supplementaires*\n")

        return "\n".join(parts)

    def _docx_to_md(self, path: Path) -> str:
        """Word -> Markdown."""
        from docx import Document

        doc = Document(str(path))
        parts = [f"# {path.name}\n"]

        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            # Detecter les titres
            if para.style.name.startswith("Heading"):
                level = int(para.style.name[-1]) if para.style.name[-1].isdigit() else 2
                parts.append(f"{'#' * (level + 1)} {text}\n")
            else:
                parts.append(text + "\n")

        return "\n".join(parts)

    def _txt_to_md(self, path: Path) -> str:
        """TXT -> Markdown (simple wrapping)."""
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = path.read_text(encoding="latin-1", errors="replace")

        return f"# {path.name}\n\n{text}\n"

    # ------------------------------------------------------------------
    # Utilitaires de nettoyage
    # ------------------------------------------------------------------

    def _clean_text(self, text: str) -> str:
        """Nettoie le texte extrait (espaces, lignes vides multiples)."""
        # Supprimer les caracteres de controle
        text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
        # Normaliser les espaces
        lines = []
        for line in text.split("\n"):
            line = line.rstrip()
            if line:
                lines.append(line)
            elif lines and lines[-1] != "":
                lines.append("")
        return "\n".join(lines)

    def _structure_fiscal_text(self, text: str) -> str:
        """Detecte les patterns fiscaux et les met en forme markdown."""
        lines = text.split("\n")
        result = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                result.append("")
                continue

            # Detecter les lignes avec des montants (pattern: texte ... montant)
            montant_match = re.search(r"(\d[\d\s]*[,\.]\d{2})\s*(?:EUR|€)?\s*$", stripped)
            if montant_match:
                # Mettre en gras le montant
                montant = montant_match.group(1)
                label = stripped[:montant_match.start()].strip().rstrip(".:;")
                if label:
                    result.append(f"- {label} : **{montant}**")
                    continue

            # Detecter les lignes qui ressemblent a des titres (tout en majuscules)
            if stripped.isupper() and len(stripped) > 3 and len(stripped) < 80:
                result.append(f"### {stripped.title()}")
                continue

            result.append(stripped)

        return "\n".join(result)

    def _ocr_bytes(self, img_bytes: bytes) -> str:
        """OCR sur des bytes d'image."""
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(io.BytesIO(img_bytes))
            return pytesseract.image_to_string(img, lang="fra")
        except Exception:
            return ""
